import openpyxl
import re
import csv

def extract_month_number(header_text):
    """从列标题中提取月份数字，如 '1月分数' -> 1, '2025年1月' -> 1"""
    match = re.search(r'(\d{1,2})\s*月', str(header_text))
    return int(match.group(1)) if match else None

def process_sheet_2026(ws, year):
    """处理2026年NPS表，表头在第1行，列名为 '1月分数'、'2月分数' 等"""
    records = []
    # 表头在第一行
    header = [cell.value for cell in ws[1]]
    month_cols = []   # (列索引, 月份数字)
    for idx, cell_val in enumerate(header):
        if cell_val and isinstance(cell_val, str) and '月分数' in cell_val:
            m = extract_month_number(cell_val)
            if m:
                month_cols.append((idx, m))

    # 数据从第2行开始
    for row in ws.iter_rows(min_row=2, values_only=True):
        company = row[1]   # B列（公司名称）
        if company is None:
            continue
        company = str(company).strip()
        if not company:
            continue
        for col_idx, month in month_cols:
            val = row[col_idx]
            if isinstance(val, (int, float)):
                records.append((company, f"{year}/{month}", float(val)))
            # 非数值忽略（如“无样本量不考核”、“NA”）
    return records

def process_sheet_2025(ws, year):
    """处理2025年多区域NPS表，扫描每个 '服务网络' 或 '区域' 子表"""
    records = []
    max_row = ws.max_row
    row = 1
    while row <= max_row:
        row_vals = [cell.value for cell in ws[row]]
        first_cell = row_vals[0] if len(row_vals) > 0 else None
        if first_cell in ('服务网络', '区域'):
            # 提取该子表的月份列（表头行可能包含“2025年1月”等）
            month_cols = []
            for idx, val in enumerate(row_vals):
                m = extract_month_number(str(val)) if val else None
                if m is not None:
                    month_cols.append((idx, m))
            # 读取数据行，直到遇到空行或下一个子表头
            data_row = row + 1
            while data_row <= max_row:
                cur_vals = [cell.value for cell in ws[data_row]]
                first = cur_vals[0] if len(cur_vals) > 0 else None
                if first is None or first in ('服务网络', '区域', '') or (isinstance(first, str) and first.strip() == ''):
                    break
                company = cur_vals[1]
                if company is not None and str(company).strip() != '':
                    company = str(company).strip()
                    for col_idx, month in month_cols:
                        val = cur_vals[col_idx]
                        if isinstance(val, (int, float)):
                            records.append((company, f"{year}/{month}", float(val)))
                data_row += 1
            row = data_row   # 跳到下一个区段
            continue
        row += 1
    return records

# ---------- 主程序 ----------
if __name__ == '__main__':
    file_2026 = r"E:\powerbi_data\看板数据\私有云文件本地\收集文件\2026年数据汇总表.xlsx"
    file_2025 = r"E:\powerbi_data\看板数据\私有云文件本地\收集文件\2025年数据汇总表.xlsx"

    all_data = []

    # 2026
    wb = openpyxl.load_workbook(file_2026, data_only=True)
    all_data.extend(process_sheet_2026(wb['NPS'], 2026))
    wb.close()

    # 2025
    wb = openpyxl.load_workbook(file_2025, data_only=True)
    all_data.extend(process_sheet_2025(wb['NPS'], 2025))
    wb.close()

    # 输出CSV
    output_file = 'E:\powerbi_data\看板数据\dashboard\merged_nps_data.csv'
    with open(output_file, 'w', newline='', encoding='utf-8-sig') as f:
        writer = csv.writer(f)
        writer.writerow(['公司名称', '月份', 'NPS分数'])
        writer.writerows(all_data)

    print(f"提取完成，共 {len(all_data)} 条记录，已保存到 {output_file}")